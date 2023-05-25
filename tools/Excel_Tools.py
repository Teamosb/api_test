# ----------------------------------------
#              创建姓名：十元

#              创建时间：2020/10/26 9:37 下午

#功能实现：
#   1：以对象的形式读取excel
#   2：按指定行和列更新单元格数据
#   3：未实现 删除单元格数据（空 是null 还是 '' 未确认）

# ---------------------------------------

import openpyxl

class ExcelData:
    def __init__(self, zipObj):
        for item in zipObj:
            # 对已存在的属性 item[0] 进行赋值 item[1]
            setattr(self, item[0], item[1])

class Exceltool:
    def __init__(self, file_name, sheet_name):
        # 初始化参数
        self.file_name = file_name
        self.sheet_name = sheet_name
        # 加载workbook和sheet
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet = self.workbook[sheet_name]
    def read_excel_obj(self):
        def func(x):
            return x.value
        # list(self.sheet_name.rows)[0] 读取第0行的数据 类型元组
        # map(func(),元组) 元组中的值依次执行func函数
        # list(map(func(),元组) ) 将第二步的值转换成list
        head = list(map(func, list(self.sheet.rows)[0]))
        datas = []
        for item in list(self.sheet.rows)[1:]:  # 从索引1到索引最后
            # 依次获取item的值 ，并转换成list
            data = list(map(func, item))
            # 依次将 头部信息(字段名)和值 打包成元组
            # 并把元组的集合转换成list
            obj = list(zip(head, data))
            # ！！！！！！！！！！！！！！应该是将list 转成对象
            excel_data = ExcelData(obj)
            # 将对象值添加进list中
            datas.append(excel_data)
        return datas

    def Excel_Write(self, row, cloume, value):
        # 依据行和列定位单元格，写入值，保存
        self.sheet.cell(row, cloume).value = value
        self.workbook.save(self.file_name)


if __name__ == '__main__':
    path = '/Users/teamo/PycharmProjects/Class02/Day17/datas/login.xlsx'
    ed = Exceltool(path, 'Sheet2').read_excel_obj()
    # 给某单元格赋值
    # cd =Excel_tools(path, 'Sheet2').Excel_Write(15,15,)
    #读取某单元格的值
    cd1 =Exceltool(path, 'Sheet2').sheet.cell(row=2,column=2).value

    print(cd1)
    print(ed)
